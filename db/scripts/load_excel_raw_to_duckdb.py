"""Excel を列名そのままで db/knowledge.duckdb に raw テーブルとして取り込む。

使い方:
    python db/scripts/load_excel_raw_to_duckdb.py raw/inbox/CASE_SAL/Case_SL12_仕様項目一覧.xlsx
    python db/scripts/load_excel_raw_to_duckdb.py <xlsx...> [--db db/knowledge.duckdb] [--prefix case_sl12]
    python db/scripts/load_excel_raw_to_duckdb.py raw/inbox/Bobcat_SAL/Bobcat_SAL_仕様値.xlsx --header-rows 2

動作:
- シートごとに `{prefix}_{シート名}` テーブルを作る（既存なら DROP して置き換え）。
- 先頭行をヘッダーとして列名をそのまま使う。列名がまちまちでもファイル単位で独立に扱う。
- --header-rows 2 で2段ヘッダーに対応。上段は結合セル分を右に引き継ぎ、`上段_下段` を列名にする
  （例: 機種名 L23 × US数値 → L23_US数値）。上段が空の列は下段のみ。
- 値は列単位で BIGINT / DOUBLE / BOOLEAN を推定し、混在列（例: 数値と 'N/A'）は原文保持のため VARCHAR にする。
- 各行に raw_path / sheet_name / row_no（Excel 上の行番号）を付け、原本へ辿れるようにする。
- 取り込み履歴を `excel_imports` テーブルに記録する。
- 取り込み後、DB 内の全テーブル定義を db/schema.sql に書き出して同期する。
"""

import argparse
import datetime as dt
import re
import sys
import unicodedata
from pathlib import Path

import duckdb
import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
META_COLUMNS = ("raw_path", "sheet_name", "row_no")


def sanitize_identifier(name: str) -> str:
    """テーブル名・列名用に空白や記号を _ に寄せる。日本語はそのまま残す。"""
    name = unicodedata.normalize("NFKC", str(name)).strip()
    name = re.sub(r"[\s/\\()\[\]{}:;,.\-]+", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    return name or "col"


def unique_names(names):
    seen = {}
    out = []
    for n in names:
        if n in seen:
            seen[n] += 1
            out.append(f"{n}_{seen[n]}")
        else:
            seen[n] = 1
            out.append(n)
    return out


def infer_column_type(values):
    """非 NULL 値から列型を推定する。混在は原文保持のため VARCHAR。"""
    non_null = [v for v in values if v is not None]
    if not non_null:
        return "VARCHAR"
    if all(isinstance(v, bool) for v in non_null):
        return "BOOLEAN"
    if all(isinstance(v, int) and not isinstance(v, bool) for v in non_null):
        return "BIGINT"
    if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in non_null):
        return "DOUBLE"
    return "VARCHAR"


def to_varchar(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, (dt.datetime, dt.date)):
        return v.isoformat()
    return str(v)


def build_header(rows, header_rows):
    """ヘッダー行群から列名リストを作る。2段以上は上段を右に引き継いで `上段_下段` に結合する。"""
    width = max(len(r) for r in rows)
    filled = [list(r) + [None] * (width - len(r)) for r in rows]
    for r in filled[:-1]:  # 最下段以外は結合セル由来の空白を直前の値で埋める
        last = None
        for i, v in enumerate(r):
            if v is not None and str(v).strip() != "":
                last = v
            else:
                r[i] = last
    header = []
    for i in range(width):
        parts = [
            str(r[i]).strip()
            for r in filled
            if r[i] is not None and str(r[i]).strip() != ""
        ]
        header.append("_".join(parts) if parts else None)
    return header


def read_sheet(ws, header_rows=1):
    """(列名リスト, データ行リスト) を返す。ヘッダー空列は捨て、全空行は飛ばす。"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < header_rows:
        return [], []
    header = build_header(rows[:header_rows], header_rows)
    keep = [i for i, h in enumerate(header) if h is not None and str(h).strip() != ""]
    if not keep:
        return [], []
    columns = unique_names([sanitize_identifier(header[i]) for i in keep])
    data = []
    for excel_row_no, row in enumerate(rows[header_rows:], start=header_rows + 1):
        values = [row[i] if i < len(row) else None for i in keep]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in values):
            continue
        data.append((excel_row_no, values))
    return columns, data


def load_workbook_to_db(con, xlsx_path: Path, prefix: str, header_rows: int = 1):
    try:
        raw_path = xlsx_path.resolve().relative_to(REPO_ROOT).as_posix()
    except ValueError:
        raw_path = xlsx_path.resolve().as_posix()

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    created = []
    for sheet_name in wb.sheetnames:
        columns, data = read_sheet(wb[sheet_name], header_rows)
        if not columns:
            print(f"  skip (empty): {sheet_name}")
            continue
        table = f"{prefix}_{sanitize_identifier(sheet_name)}"

        col_values = list(zip(*[values for _, values in data])) if data else [[]] * len(columns)
        types = [infer_column_type(list(vs)) for vs in col_values]

        col_defs = ", ".join(
            ['"raw_path" VARCHAR', '"sheet_name" VARCHAR', '"row_no" BIGINT']
            + [f'"{c}" {t}' for c, t in zip(columns, types)]
        )
        con.execute(f'DROP TABLE IF EXISTS "{table}"')
        con.execute(f'CREATE TABLE "{table}" ({col_defs})')

        insert_rows = []
        for excel_row_no, values in data:
            converted = [
                to_varchar(v) if t == "VARCHAR" else v
                for v, t in zip(values, types)
            ]
            insert_rows.append([raw_path, sheet_name, excel_row_no] + converted)
        if insert_rows:
            placeholders = ", ".join(["?"] * (len(columns) + len(META_COLUMNS)))
            con.executemany(f'INSERT INTO "{table}" VALUES ({placeholders})', insert_rows)

        con.execute(
            "INSERT INTO excel_imports VALUES (?, ?, ?, ?, ?, current_timestamp)",
            [table, raw_path, sheet_name, ", ".join(columns), len(insert_rows)],
        )
        created.append((table, len(insert_rows)))
        print(f"  {table}: {len(insert_rows)} rows ({', '.join(columns)})")
    return created


def drop_existing_prefix_tables(con, prefix: str):
    """同じファイル由来の旧テーブルを先に落とす。削除済みシートの残骸を残さないため。"""
    tables = con.execute(
        "SELECT table_name FROM information_schema.tables "
        "WHERE table_schema = 'main' AND starts_with(table_name, ?) "
        "ORDER BY table_name",
        [prefix + "_"],
    ).fetchall()
    for (table_name,) in tables:
        con.execute(f'DROP TABLE IF EXISTS "{table_name.replace(chr(34), chr(34) + chr(34))}"')


def dump_schema(con, schema_path: Path):
    rows = con.execute(
        "SELECT sql FROM duckdb_tables() WHERE NOT internal ORDER BY table_name"
    ).fetchall()
    views = con.execute(
        "SELECT sql FROM duckdb_views() WHERE NOT internal ORDER BY view_name"
    ).fetchall()
    lines = [
        "-- db/schema.sql",
        "-- db/scripts/load_excel_raw_to_duckdb.py が knowledge.duckdb の実テーブルから自動生成。",
        f"-- 最終更新: {dt.date.today().isoformat()}",
        "",
    ]
    for (sql,) in rows + views:
        sql = sql.rstrip().rstrip(";")
        lines.append(sql + ";")
        lines.append("")
    schema_path.write_text("\n".join(lines), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("files", nargs="+", help="取り込む xlsx ファイル")
    parser.add_argument("--db", default=str(REPO_ROOT / "db" / "knowledge.duckdb"))
    parser.add_argument(
        "--prefix",
        default=None,
        help="テーブル名プレフィックス（省略時はファイル名から生成。複数ファイル時は無視）",
    )
    parser.add_argument(
        "--header-rows",
        type=int,
        default=1,
        help="ヘッダー行数（既定 1。2 なら上段を右に引き継いで `上段_下段` に結合）",
    )
    args = parser.parse_args()

    con = duckdb.connect(args.db)
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS excel_imports (
            table_name VARCHAR,
            raw_path VARCHAR,
            sheet_name VARCHAR,
            columns VARCHAR,
            row_count BIGINT,
            loaded_at TIMESTAMP
        )
        """
    )

    for f in args.files:
        xlsx_path = Path(f)
        if not xlsx_path.exists():
            sys.exit(f"not found: {f}")
        if args.prefix and len(args.files) == 1:
            prefix = args.prefix
        else:
            prefix = sanitize_identifier(xlsx_path.stem).lower()
        print(f"{xlsx_path.name} -> prefix '{prefix}'")
        con.execute("DELETE FROM excel_imports WHERE starts_with(table_name, ?)", [prefix + "_"])
        drop_existing_prefix_tables(con, prefix)
        load_workbook_to_db(con, xlsx_path, prefix, args.header_rows)

    dump_schema(con, REPO_ROOT / "db" / "schema.sql")
    con.close()
    print(f"done. db={args.db}, schema=db/schema.sql")


if __name__ == "__main__":
    main()
