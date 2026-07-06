from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def sanitize_sheet_name(sheet_name: str) -> str:
    return "".join(character if character.isalnum() or character in {"-", "_"} else "_" for character in sheet_name).strip("_") or "sheet"


def write_sheet_csv(sheet, output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert an Excel workbook to CSV files.")
    parser.add_argument("xlsx", type=Path, help="Path to the source .xlsx file")
    parser.add_argument("--output-dir", type=Path, help="Directory to write CSV files to. Defaults to raw/data.")
    parser.add_argument("--sheet", help="Optional sheet name to convert. Defaults to all sheets.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing CSV files.")
    args = parser.parse_args()

    xlsx_path = args.xlsx
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    repo_root = Path(__file__).resolve().parents[1]
    output_dir = args.output_dir or repo_root / "raw" / "data"
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet_names = [args.sheet] if args.sheet else workbook.sheetnames
        missing_sheets = [sheet_name for sheet_name in sheet_names if sheet_name not in workbook.sheetnames]
        if missing_sheets:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(f"Sheet not found: {missing_sheets[0]} (available: {available})")

        multiple_sheets = len(sheet_names) > 1
        multi_sheet_dir = output_dir / xlsx_path.stem if multiple_sheets else output_dir
        multi_sheet_dir.mkdir(parents=True, exist_ok=True)
        written_paths: list[Path] = []
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            output_name = f"{sanitize_sheet_name(sheet_name)}.csv" if multiple_sheets else f"{xlsx_path.stem}.csv"
            output_path = multi_sheet_dir / output_name
            if output_path.exists() and not args.overwrite:
                raise FileExistsError(f"CSV file already exists: {output_path}")
            write_sheet_csv(sheet, output_path)
            written_paths.append(output_path)
    finally:
        workbook.close()

    for written_path in written_paths:
        print(written_path)


if __name__ == "__main__":
    main()