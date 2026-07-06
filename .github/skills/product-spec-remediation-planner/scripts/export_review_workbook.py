# -*- coding: utf-8 -*-
"""人間レビュー用の Excel・HTMLダッシュボード・決定テンプレートCSVを作る。

使い方:
    python .github/skills/product-spec-remediation-planner/scripts/export_review_workbook.py
        [--audit-dir audit_outputs] [--out-dir remediation_outputs]

前提:
    scripts/analyze_audit_outputs.py の実行済み（未実行なら自動で実行する）。

出力（remediation_outputs/ 配下）:
    - human_review_workbook.xlsx     人間レビュー用ワークブック（1行1判断・白地黒文字・フィルタ不使用）
    - human_review_required.csv      人間レビューが必要な項目（review_id とシート位置付き）
    - review_dashboard.html          読み取り専用サマリ（判断入力用ではない）
    - reviewed_decision_template.csv レビュー結果を機械で読める形で渡すためのテンプレート

動作:
    - DB には接続しない。normalizer のファイルにも触れない。
    - review_id を全シート・全CSVの共通キーにする。
    - 01_priority_review には P1 の行を集約再掲する（専門シートと同じ review_id）。
      両方に異なる判断が記入された場合は import 時に不整合として弾かれる。
"""
import argparse
import html
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from analyze_audit_outputs import (
    DECISION_VOCAB, STATUS_VOCAB, DEFAULT_AUDIT_DIR, DEFAULT_OUT_DIR,
    ensure_analysis, write_csv,
)

# 判断シートの列（human_decision 以降が人間の記入欄）
REVIEW_COLS = [
    "review_id", "priority", "issue_type", "question", "ai_recommendation",
    "ai_uncertainty", "decision_options", "confidence", "affected_count",
    "sample_values", "source_examples",
    "human_decision", "final_canonical_item", "final_unit",
    "apply_flag", "reviewer_note", "status",
]
HUMAN_COLS = {"human_decision", "final_canonical_item", "final_unit",
              "apply_flag", "reviewer_note", "status"}

COL_WIDTHS = {
    "review_id": 11, "priority": 8, "issue_type": 22, "question": 46,
    "ai_recommendation": 40, "ai_uncertainty": 34, "decision_options": 30,
    "confidence": 10, "affected_count": 11, "sample_values": 32,
    "source_examples": 30, "human_decision": 22, "final_canonical_item": 24,
    "final_unit": 10, "apply_flag": 10, "reviewer_note": 30, "status": 10,
}

SHEET_BY_TYPE = {
    "risky_group_merge": "02_risky_group_merges",
    "unmapped_raw_item": "03_unmapped_raw_items",
    "low_confidence_mapping": "03_unmapped_raw_items",
    "unit_mismatch": "04_unit_mismatches",
    "duplicate_canonical_item": "05_duplicate_items",
    "canonical_split_proposal": "06_canonical_split_proposals",
    "exclusion_candidate": "07_exclusion_candidates",
}

README_LINES = [
    "human_review_workbook.xlsx — 記入手順",
    "",
    "1. 01_priority_review を上から順に判断する（P1 = 比較結果を誤らせる恐れのある問題）。",
    "2. 時間があれば 02〜07 の各シートで残りを判断する。",
    "3. 記入するのは human_decision / final_canonical_item / final_unit / apply_flag /",
    "   reviewer_note / status の6列だけ。他の列と review_id は編集しない。",
    "4. human_decision はドロップダウンの固定語彙から選ぶ（自由記述は取り込み時にエラー）。",
    "5. 判断したら status を reviewed にする。後回しは deferred。",
    "6. map_to_existing / create_new_canonical / split_canonical → final_canonical_item 必須。",
    "   add_conversion_rule / change_canonical_unit → final_unit 必須。",
    "7. 修正に反映してよい行だけ apply_flag を true にする。",
    "8. reject / needs_source_check の場合は reviewer_note に理由を必ず書く。",
    "",
    "注意: 01_priority_review の行は専門シート（02〜06）にも同じ review_id で再掲されている。",
    "どちらに記入してもよいが、両方に異なる判断を書くと不整合として弾かれる。",
    "",
    "ai_recommendation は AI の候補であり従う義務はない。人間判断が常に優先される。",
    "迷ったら needs_source_check / keep_manual_review にする（誤った approve の方が害が大きい）。",
    "",
    "詳細: .github/skills/product-spec-remediation-planner/docs/human_review_guide.md",
]

THIN = Border(*(Side(style="thin", color="000000"),) * 4)


def finding_to_row(f):
    return {
        "review_id": f["review_id"], "priority": f["priority"],
        "issue_type": f["issue_type"], "question": f["question"],
        "ai_recommendation": f["ai_recommendation"],
        "ai_uncertainty": f["ai_uncertainty"],
        "decision_options": " / ".join(f["decision_options"]),
        "confidence": f["confidence"], "affected_count": f["affected_count"],
        "sample_values": f["sample_values"], "source_examples": f["source_examples"],
        "human_decision": "", "final_canonical_item": "", "final_unit": "",
        "apply_flag": "", "reviewer_note": "", "status": "pending",
    }


def write_review_sheet(wb, title, rows, note=""):
    ws = wb.create_sheet(title)
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for c, name in enumerate(REVIEW_COLS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.border = THIN
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS.get(name, 14)
    r = 1
    for r, row in enumerate(rows, 2):
        for c, name in enumerate(REVIEW_COLS, 1):
            cell = ws.cell(row=r, column=c, value=row.get(name, ""))
            cell.alignment = wrap
            cell.border = THIN
    if not rows and note:
        ws.cell(row=2, column=1, value=note)
    ws.freeze_panes = "D2"
    if rows:
        last = len(rows) + 1
        col = {name: get_column_letter(i + 1) for i, name in enumerate(REVIEW_COLS)}
        dv_decision = DataValidation(
            type="list", formula1='"' + ",".join(DECISION_VOCAB) + '"', allow_blank=True,
            showErrorMessage=True, errorTitle="固定語彙のみ",
            error="human_decision は固定語彙から選んでください")
        dv_status = DataValidation(
            type="list", formula1='"' + ",".join(STATUS_VOCAB) + '"', allow_blank=True)
        dv_flag = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
        for dv, name in ((dv_decision, "human_decision"), (dv_status, "status"),
                         (dv_flag, "apply_flag")):
            ws.add_data_validation(dv)
            dv.add(f"{col[name]}2:{col[name]}{last}")
    return ws


def build_workbook(analysis, out_path: Path):
    findings = sorted(analysis["findings"],
                      key=lambda f: (f["priority"], -f["affected_count"]))
    wb = Workbook()
    ws = wb.active
    ws.title = "00_readme"
    ws.column_dimensions["A"].width = 100
    for i, line in enumerate(README_LINES, 1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=12)

    p1_rows = [finding_to_row(f) for f in findings if f["priority"] == "P1"]
    write_review_sheet(wb, "01_priority_review", p1_rows, "P1項目なし")

    sheet_rows = {name: [] for name in [
        "02_risky_group_merges", "03_unmapped_raw_items", "04_unit_mismatches",
        "05_duplicate_items", "06_canonical_split_proposals", "07_exclusion_candidates"]}
    for f in findings:
        sheet = SHEET_BY_TYPE.get(f["issue_type"])
        if sheet:
            sheet_rows[sheet].append(finding_to_row(f))
    for name, rows in sheet_rows.items():
        write_review_sheet(wb, name, rows, "該当なし")

    log = wb.create_sheet("08_decision_log")
    for c, name in enumerate(["date", "reviewer", "review_id", "action", "note"], 1):
        log.cell(row=1, column=c, value=name).font = Font(bold=True)
        log.column_dimensions[get_column_letter(c)].width = [12, 14, 12, 30, 50][c - 1]
    log.cell(row=2, column=1, value="(レビューのやり直し・差し戻し等の経緯を自由記入)")

    wb.save(out_path)
    return {name: len(rows) for name, rows in
            [("01_priority_review", p1_rows)] + list(sheet_rows.items())}


def build_dashboard(analysis, sheet_counts, out_path: Path):
    m = analysis["current_metrics"]
    findings = analysis["findings"]
    esc = html.escape

    def fmt(v):
        return "不明" if v is None else (f"{v:g}" if isinstance(v, float) else str(v))

    prio_counts = {}
    for f in findings:
        prio_counts[f["priority"]] = prio_counts.get(f["priority"], 0) + 1
    n_human = sum(1 for f in findings if f["needs_human"])
    risky = [f for f in findings if f["issue_type"] == "risky_group_merge"]
    p1 = [f for f in findings if f["priority"] == "P1"]

    def table_html(headers, rows):
        head = "".join(f"<th>{esc(str(h))}</th>" for h in headers)
        body = "".join(
            "<tr>" + "".join(f"<td>{esc(str(c))}</td>" for c in row) + "</tr>" for row in rows)
        return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"

    metric_cards = "".join(
        f"<div class='card'><div class='num'>{esc(fmt(v))}</div><div class='lbl'>{esc(k)}</div></div>"
        for k, v in [
            ("総合スコア /100", m.get("total_score")),
            ("未マッピング率 %", m.get("unmapped_pct")),
            ("単位不一致 行", m.get("unit_mismatch_rows")),
            ("値競合 組", m.get("value_conflicts")),
            ("人間レビュー件数", n_human),
        ])

    sheet_guide = [
        ("01_priority_review", "P1（比較を誤らせる恐れ）の全件。まずここだけでも完了させる"),
        ("02_risky_group_merges", "意味混同疑いの canonical 群"),
        ("03_unmapped_raw_items", "未マッピング項目と低confidenceマッピング"),
        ("04_unit_mismatches", "単位不一致（変換ルール追加/変換禁止の判断）"),
        ("05_duplicate_items", "同一製品×項目の重複"),
        ("06_canonical_split_proposals", "canonical_item 分割案の承認"),
        ("07_exclusion_candidates", "比較対象外候補"),
    ]
    doc = f"""<!DOCTYPE html>
<html lang="ja"><head><meta charset="utf-8">
<title>remediation review dashboard</title>
<style>
 body {{ font-family: "Yu Gothic UI", Meiryo, sans-serif; background:#fff; color:#000;
        margin: 24px; line-height: 1.5; }}
 h1 {{ font-size: 20px; border-bottom: 2px solid #000; padding-bottom: 6px; }}
 h2 {{ font-size: 16px; margin-top: 28px; }}
 table {{ border-collapse: collapse; margin: 8px 0; font-size: 13px; }}
 th, td {{ border: 1px solid #000; padding: 4px 8px; text-align: left; vertical-align: top; }}
 th {{ background: #f0f0f0; }}
 .cards {{ display: flex; gap: 12px; flex-wrap: wrap; margin: 12px 0; }}
 .card {{ border: 1px solid #000; padding: 10px 16px; min-width: 120px; text-align:center; }}
 .num {{ font-size: 24px; font-weight: bold; }}
 .lbl {{ font-size: 12px; }}
 .note {{ font-size: 12px; color: #333; }}
</style></head><body>
<h1>product-spec-normalizer 修正レビュー ダッシュボード</h1>
<p class="note">生成: {esc(analysis['generated_at'])} ／ このページは読み取り専用のサマリです。
判断の記入は human_review_workbook.xlsx で行ってください。</p>
<div class="cards">{metric_cards}</div>
<h2>優先度別件数</h2>
{table_html(["優先度", "判断件数"], sorted(prio_counts.items()))}
<h2>重大問題（P1）一覧</h2>
{table_html(["review_id", "種別", "問題", "影響件数"],
            [[f['review_id'], f['issue_type'], f['question'], f['affected_count']] for f in p1])}
<h2>risky group 一覧（意味混同疑い）</h2>
{table_html(["review_id", "group", "canonical_item", "分割案", "影響行"],
            [[f['review_id'], f.get('rule_name', ''), f.get('canonical_item', ''),
              f['ai_recommendation'], f['affected_count']] for f in risky])}
<h2>どのシートを見ればよいか</h2>
{table_html(["シート", "内容", "行数"],
            [[name, desc, sheet_counts.get(name, 0)] for name, desc in sheet_guide])}
<p class="note">不足していた監査入力: {esc(', '.join(analysis['missing_inputs']) or 'なし')}
／ risky情報源: {esc(analysis['risky_source'])}</p>
</body></html>"""
    out_path.write_text(doc, encoding="utf-8")


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--audit-dir", type=Path, default=DEFAULT_AUDIT_DIR)
    ap.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = ap.parse_args(argv)
    analysis = ensure_analysis(args.audit_dir, args.out_dir)
    findings = analysis["findings"]

    wb_path = args.out_dir / "human_review_workbook.xlsx"
    sheet_counts = build_workbook(analysis, wb_path)

    manual = sorted([f for f in findings if f["needs_human"]],
                    key=lambda f: (f["priority"], -f["affected_count"]))
    write_csv(args.out_dir / "human_review_required.csv",
              ["review_id", "priority", "sheet", "issue_type", "question",
               "ai_recommendation", "affected_count", "confidence"],
              [{**f, "sheet": ("01_priority_review" if f["priority"] == "P1"
                               else SHEET_BY_TYPE.get(f["issue_type"], ""))} for f in manual])

    write_csv(args.out_dir / "reviewed_decision_template.csv",
              ["review_id", "issue_type", "human_decision", "final_canonical_item",
               "final_unit", "apply_flag", "reviewer_note", "reviewed_by", "reviewed_at"],
              [{"review_id": f["review_id"], "issue_type": f["issue_type"]}
               for f in sorted(findings, key=lambda x: (x["priority"], -x["affected_count"]))])

    build_dashboard(analysis, sheet_counts, args.out_dir / "review_dashboard.html")

    print(f"written: {wb_path}")
    for name, n in sheet_counts.items():
        print(f"  {name}: {n}行")
    print(f"written: {args.out_dir / 'human_review_required.csv'} ({len(manual)}件)")
    print(f"written: {args.out_dir / 'reviewed_decision_template.csv'}")
    print(f"written: {args.out_dir / 'review_dashboard.html'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
