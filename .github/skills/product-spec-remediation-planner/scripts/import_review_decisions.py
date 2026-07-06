# -*- coding: utf-8 -*-
"""人間レビュー済みの Excel または CSV を読み込み、検証して承認済みCSVに分ける。

使い方:
    python .github/skills/product-spec-remediation-planner/scripts/import_review_decisions.py
        [--out-dir remediation_outputs] [--decisions <xlsx|csv>]

入力（--decisions 省略時は次の優先順で自動選択）:
    1. remediation_outputs/reviewed_decisions.csv
    2. remediation_outputs/human_review_workbook.xlsx
    3. remediation_outputs/reviewed_decision_template.csv（human_decision が1件でも記入済みの場合）

出力（remediation_outputs/ 配下）:
    - approved_mapping_changes.csv    承認済みマッピング修正（map_to_existing / approve /
                                      fix_raw_unit_extraction 等。change_type 列で区別）
    - approved_canonical_splits.csv   承認済み canonical 分割
    - approved_new_canonical_items.csv 承認済み新規 canonical
    - approved_exclusions.csv         承認済み比較対象外
    - approved_unit_rules.csv         承認済み単位変換ルール（add_conversion_rule /
                                      change_canonical_unit）
    - conversion_blocklist.csv        承認済み変換禁止（do_not_convert）
    - manual_review_remaining.csv     修正対象に入れない行（保留・差し戻し・reject 等）
    - reviewed_decision_errors.csv    不備行（語彙外・必須欠落・不整合）
    - _decisions.json                 build_normalizer_change_plan.py 用の内部データ

検証・分類規則（SKILL.md「human_decision の扱い」と一致）:
    - human_decision は固定語彙のみ。status は pending / reviewed / deferred のみ。
    - approved に入る条件: status=reviewed かつ apply_flag=true かつ human_decision が
      reject / needs_source_check / keep_manual_review 以外で、必須項目が揃っていること。
    - 同一 review_id が複数シートで異なる判断を持つ場合は inconsistent_decision。
    - 人間判断とAI推奨が矛盾しても常に人間判断を採用する（AI推奨は記録のみ）。

動作:
    - DB には接続しない。normalizer のファイルにも触れない。
"""
import argparse
import csv
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from analyze_audit_outputs import (
    DECISIONS_REQUIRE_CANONICAL, DECISIONS_REQUIRE_UNIT, DECISION_VOCAB,
    DEFAULT_OUT_DIR, STATUS_VOCAB, load_analysis, write_csv,
)

DECISION_COLS = ["review_id", "issue_type", "human_decision", "final_canonical_item",
                 "final_unit", "apply_flag", "reviewer_note", "reviewed_by", "reviewed_at",
                 "status", "source_sheet"]

REJECT_LIKE = {"reject", "needs_source_check", "keep_manual_review"}
TRUE_WORDS = {"true", "1", "yes", "y"}
FALSE_WORDS = {"false", "0", "no", "n", ""}


def norm(v):
    return str(v).strip() if v is not None else ""


def read_decisions_csv(path: Path):
    with path.open(encoding="utf-8-sig", newline="") as f:
        rows = []
        for r in csv.DictReader(f):
            row = {c: norm(r.get(c)) for c in DECISION_COLS}
            row["source_sheet"] = path.name
            # CSV には status 列が無い場合がある: 記入済みなら reviewed とみなす
            if not row["status"]:
                row["status"] = "reviewed" if row["human_decision"] else "pending"
            rows.append(row)
        return rows


def read_decisions_xlsx(path: Path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []
    for name in wb.sheetnames:
        if not (name[:2].isdigit() and name[:2] not in ("00", "08")):
            continue
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            continue
        idx = {str(h): i for i, h in enumerate(header) if h}
        if "review_id" not in idx or "human_decision" not in idx:
            continue
        for values in it:
            def get(colname):
                i = idx.get(colname)
                return norm(values[i]) if i is not None and i < len(values) else ""
            rid = get("review_id")
            if not rid or not rid[:3].isalpha():
                continue
            rows.append({
                "review_id": rid, "issue_type": get("issue_type"),
                "human_decision": get("human_decision"),
                "final_canonical_item": get("final_canonical_item"),
                "final_unit": get("final_unit"),
                "apply_flag": get("apply_flag"), "reviewer_note": get("reviewer_note"),
                "reviewed_by": get("reviewed_by"), "reviewed_at": get("reviewed_at"),
                "status": get("status") or "pending", "source_sheet": name,
            })
    wb.close()
    return rows


def is_filled(row):
    return bool(row["human_decision"]) or row["status"] in ("reviewed", "deferred")


def merge_by_review_id(rows):
    """複数シート（01再掲と専門シート）の行を review_id 単位に統合する。

    戻り値: (統合後の行リスト, 不整合review_idの集合)
    """
    groups = defaultdict(list)
    for r in rows:
        groups[r["review_id"]].append(r)
    merged, inconsistent = [], set()
    keys = ("human_decision", "final_canonical_item", "final_unit", "apply_flag", "status")
    for rid, members in groups.items():
        filled = [m for m in members if is_filled(m)]
        if not filled:
            merged.append(members[0])
            continue
        signatures = {tuple(m[k].lower() for k in keys) for m in filled}
        if len(signatures) > 1:
            inconsistent.add(rid)
            merged.append(filled[0])
        else:
            merged.append(filled[0])
    return merged, inconsistent


def classify_row(row, inconsistent_ids):
    """行を approved / rejected / needs_source_check / deferred /
    inconsistent_decision / missing_required_field / pending / error に分類する。
    戻り値: (分類, エラー理由リスト)
    """
    errors = []
    decision = row["human_decision"]
    status = row["status"]
    if row["review_id"] in inconsistent_ids:
        return "inconsistent_decision", ["複数シートで異なる判断が記入されている"]
    if status and status not in STATUS_VOCAB:
        errors.append(f"status が固定語彙外: {status}")
    if decision and decision not in DECISION_VOCAB:
        errors.append(f"human_decision が固定語彙外: {decision}")
    if errors:
        return "error", errors
    if not decision or status == "pending":
        return "pending", []
    if status == "deferred":
        return "deferred", []
    if decision == "reject":
        return "rejected", []
    if decision == "needs_source_check":
        return "needs_source_check", []
    if decision == "keep_manual_review":
        return "deferred", []
    if decision in DECISIONS_REQUIRE_CANONICAL and not row["final_canonical_item"]:
        return "missing_required_field", [f"{decision} には final_canonical_item が必須"]
    if decision in DECISIONS_REQUIRE_UNIT and not row["final_unit"]:
        return "missing_required_field", [f"{decision} には final_unit が必須"]
    flag = row["apply_flag"].lower()
    if flag in TRUE_WORDS:
        return "approved", []
    if flag in FALSE_WORDS:
        return "not_applied", []
    return "error", [f"apply_flag が true/false でない: {row['apply_flag']}"]


def pick_input(out_dir: Path, explicit: Path):
    if explicit:
        return explicit
    csv_path = out_dir / "reviewed_decisions.csv"
    if csv_path.exists():
        return csv_path
    xlsx_path = out_dir / "human_review_workbook.xlsx"
    if xlsx_path.exists():
        return xlsx_path
    template = out_dir / "reviewed_decision_template.csv"
    if template.exists():
        rows = read_decisions_csv(template)
        if any(r["human_decision"] for r in rows):
            return template
    return None


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    ap.add_argument("--decisions", type=Path, default=None,
                    help="レビュー済み xlsx または csv のパス")
    args = ap.parse_args(argv)
    out_dir = args.out_dir

    src = pick_input(out_dir, args.decisions)
    if src is None:
        print("エラー: レビュー済みの入力が見つからない。", file=sys.stderr)
        print("  reviewed_decisions.csv / human_review_workbook.xlsx のどちらかを "
              f"{out_dir} に置くか、--decisions で指定する。", file=sys.stderr)
        return 1
    print(f"入力: {src}")
    rows = (read_decisions_xlsx(src) if src.suffix.lower() == ".xlsx"
            else read_decisions_csv(src))
    merged, inconsistent = merge_by_review_id(rows)

    # AI分析（issue詳細）と結合して承認済みCSVに文脈を付ける
    analysis = load_analysis(out_dir)
    meta = {f["review_id"]: f for f in (analysis or {}).get("findings", [])}

    buckets = defaultdict(list)
    error_rows = []
    for row in merged:
        cls, why = classify_row(row, inconsistent)
        f = meta.get(row["review_id"], {})
        joined = {**row,
                  "issue_type": row["issue_type"] or f.get("issue_type", ""),
                  "ai_recommendation": f.get("ai_recommendation", ""),
                  "recommended_decision": f.get("recommended_decision", ""),
                  "priority": f.get("priority", ""),
                  "affected_count": f.get("affected_count", ""),
                  "canonical_item": f.get("canonical_item", ""),
                  "raw_item_name": f.get("raw_item_name", ""),
                  "current_canonical_item": f.get("current_canonical_item", ""),
                  "target_raw_items": f.get("target_raw_items", ""),
                  "original_unit": f.get("original_unit", ""),
                  "target_unit": f.get("target_unit", ""),
                  "conversion_rule": f.get("conversion_rule", ""),
                  "proposed_canonical": f.get("proposed_canonical", ""),
                  "classification": cls, "error_reason": "; ".join(why)}
        if not f and row["review_id"]:
            joined["error_reason"] = (joined["error_reason"] + "; " if why else "") + \
                "review_id が _analysis.json に見つからない（IDの編集ミスの疑い）"
            if cls == "approved":
                cls = joined["classification"] = "error"
        if cls in ("error", "inconsistent_decision", "missing_required_field"):
            error_rows.append(joined)
            if cls != "error":
                buckets[cls].append(joined)
        else:
            buckets[cls].append(joined)

    approved = buckets["approved"]

    def sel(rows_, decisions):
        return [r for r in rows_ if r["human_decision"] in decisions]

    write_csv(out_dir / "approved_canonical_splits.csv",
              ["review_id", "issue_type", "current_canonical_item", "final_canonical_item",
               "target_raw_items", "reviewer_note", "reviewed_by"],
              [{**r, "current_canonical_item":
                r["current_canonical_item"] or r["canonical_item"]}
               for r in sel(approved, {"split_canonical"})])

    write_csv(out_dir / "approved_new_canonical_items.csv",
              ["review_id", "raw_item_name", "final_canonical_item", "proposed_canonical",
               "reviewer_note", "reviewed_by"],
              sel(approved, {"create_new_canonical"}))

    write_csv(out_dir / "approved_exclusions.csv",
              ["review_id", "issue_type", "raw_item_name", "reviewer_note", "reviewed_by"],
              sel(approved, {"exclude_from_comparison"}))

    write_csv(out_dir / "approved_unit_rules.csv",
              ["review_id", "canonical_item", "original_unit", "target_unit", "final_unit",
               "conversion_rule", "human_decision", "reviewer_note", "reviewed_by"],
              sel(approved, {"add_conversion_rule", "change_canonical_unit"}))

    write_csv(out_dir / "conversion_blocklist.csv",
              ["review_id", "canonical_item", "original_unit", "target_unit",
               "reviewer_note", "reviewed_by"],
              sel(approved, {"do_not_convert"}))

    mapping_changes = []
    for r in approved:
        d = r["human_decision"]
        if d in {"map_to_existing"}:
            change = "remap"
        elif d == "approve":
            change = {"duplicate_canonical_item": "deduplicate",
                      "low_confidence_mapping": "confirm_mapping"}.get(
                          r["issue_type"], "apply_ai_recommendation")
        elif d == "fix_raw_unit_extraction":
            change = "fix_raw_unit_extraction"
        else:
            continue
        mapping_changes.append({**r, "change_type": change})
    write_csv(out_dir / "approved_mapping_changes.csv",
              ["review_id", "issue_type", "change_type", "raw_item_name", "canonical_item",
               "final_canonical_item", "ai_recommendation", "reviewer_note", "reviewed_by"],
              mapping_changes)

    remaining = (buckets["pending"] + buckets["deferred"] + buckets["rejected"] +
                 buckets["needs_source_check"] + buckets["not_applied"] +
                 buckets["inconsistent_decision"] + buckets["missing_required_field"])
    write_csv(out_dir / "manual_review_remaining.csv",
              ["review_id", "issue_type", "priority", "classification", "human_decision",
               "status", "apply_flag", "reviewer_note", "error_reason"],
              sorted(remaining, key=lambda r: (r["priority"] or "P9", r["review_id"])))

    write_csv(out_dir / "reviewed_decision_errors.csv",
              ["review_id", "source_sheet", "classification", "error_reason",
               "human_decision", "status", "apply_flag", "final_canonical_item", "final_unit"],
              error_rows)

    decisions_meta = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": str(src),
        "counts": {k: len(v) for k, v in buckets.items()},
        "n_errors": len(error_rows),
        "approved": approved,
        "rejected": buckets["rejected"],
        "remaining": remaining,
        "mapping_changes": mapping_changes,
    }
    with (out_dir / "_decisions.json").open("w", encoding="utf-8") as fh:
        json.dump(decisions_meta, fh, ensure_ascii=False, indent=1)

    print(f"レビュー行: {len(rows)}（統合後 {len(merged)}）")
    for k in ("approved", "rejected", "needs_source_check", "deferred", "pending",
              "not_applied", "inconsistent_decision", "missing_required_field"):
        if buckets[k]:
            print(f"  {k}: {len(buckets[k])}")
    print(f"不備行: {len(error_rows)} → reviewed_decision_errors.csv")
    print(f"修正対象外: {len(remaining)} → manual_review_remaining.csv")
    return 0


if __name__ == "__main__":
    sys.exit(main())
